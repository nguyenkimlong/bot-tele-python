#!/usr/bin/env python
# pylint: disable=unused-argument
# This program is dedicated to the public domain under the CC0 license.

"""
Don't forget to enable inline mode with @BotFather

First, a few handler functions are defined. Then, those functions are passed to
the Application and registered at their respective places.
Then, the bot is started and runs until we press Ctrl-C on the command line.

Usage:
Basic inline bot example. Applies different text transformations.
Press Ctrl-C on the command line or send a signal to the process to stop the
bot.
"""
import logging
from html import escape
from uuid import uuid4
import os

from telegram import InlineQueryResultArticle, InputTextMessageContent, Update
from telegram.constants import ParseMode,ChatAction
from telegram.ext import Application, CommandHandler, ContextTypes, InlineQueryHandler,CallbackContext
import sqlite3 as lite

import pandas as pd
from openpyxl import load_workbook


from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

import io
import fitz 

from PIL import Image
# Enable logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
# set higher logging level for httpx to avoid all GET and POST requests being logged
logging.getLogger("httpx").setLevel(logging.WARNING)

logger = logging.getLogger(__name__)

database_path = os.path.dirname(__file__) + "\\MyData.db"
# cur = con.cursor()    
# cur.execute('SELECT SQLITE_VERSION()')
     
# data = cur.fetchone()
     
# print ("SQLite version: %s" % data) 

class Employee:
    id = 0
    Code = ""
    phone= ""
    EmployeeFullName = ""
    EmployeeOnlyName = ""



def remove_sign_for_vietnamese_string(str=""):
    if not str:
        return str
    
    vietnamese_signs  = [        

            "aAeEoOuUiIdDyY",

            "áàạảãâấầậẩẫăắằặẳẵ",

            "ÁÀẠẢÃÂẤẦẬẨẪĂẮẰẶẲẴ",

            "éèẹẻẽêếềệểễ",

            "ÉÈẸẺẼÊẾỀỆỂỄ",

            "óòọỏõôốồộổỗơớờợởỡ",

            "ÓÒỌỎÕÔỐỒỘỔỖƠỚỜỢỞỠ",

            "úùụủũưứừựửữ",

            "ÚÙỤỦŨƯỨỪỰỬỮ",

            "íìịỉĩ",

            "ÍÌỊỈĨ",

            "đ",

            "Đ",

            "ýỳỵỷỹ",

            "ÝỲỴỶỸ"
        ]

    for i in range(1, len(vietnamese_signs)):
        for j in range(len(vietnamese_signs[i])):
            str = str.replace(vietnamese_signs[i][j], vietnamese_signs[0][i - 1])

    return str



def find_employee_by_code(name):
    name = None

    # Connect to the SQLite database
    conn = lite.connect(database_path)
    cursor = conn.cursor()

    try:
        # Remove Vietnamese signs from clone
        str = remove_sign_for_vietnamese_string(name)

        # Find an employee with a matching code (case-insensitive)
        cursor.execute("SELECT * FROM Employees WHERE lower(Code) = ?", (str.lower()))
        any_code = cursor.fetchone()

        if any_code is None:
            # Send a message and break if no employee is found
            name = ""
        else:
            # Remove Vietnamese signs from the employee's full name and replace spaces with underscores
            name = remove_sign_for_vietnamese_string(any_code[1]).replace(" ", "_")

    finally:
        # Close the database connection
        conn.close()

    return name
        


# Define a few command handlers. These usually take the two arguments update and
# context.
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Send a message when the command /start is issued."""
    helloMess = f"Xin chào, {update.message.from_user}! \n Anh chị gõ lệnh: \n /help \n Để biết cách lấy chấm công"
    await update.message.reply_text(helloMess)


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Send a message when the command /help is issued."""
    des = "Hướng dẫn lấy ngày chấm công:\n Ví dụ a/c có mã số là NVA1 thì câu lệnh sẽ là: \n/msnv NVA1"
    await update.message.reply_text(des)

async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    args = context.args
    commandParameters = args[0]
    await update.message.reply_chat_action(action=ChatAction.TYPING)
    """Send a message when the command /help is issued."""
    if commandParameters != "" and commandParameters == "initDB":
        mess = excel_to_db()
     
        await update.message.reply_text(mess)                  
                    


async def inline_query(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle the inline query. This is run when you type: @botusername <query>"""
    query = update.inline_query.query

    if not query:  # empty query should not be handled
        return

    results = [
        InlineQueryResultArticle(
            id=str(uuid4()),
            title="Caps",
            input_message_content=InputTextMessageContent(query.upper()),
        ),
        InlineQueryResultArticle(
            id=str(uuid4()),
            title="Bold",
            input_message_content=InputTextMessageContent(
                f"<b>{escape(query)}</b>", parse_mode=ParseMode.HTML
            ),
        ),
        InlineQueryResultArticle(
            id=str(uuid4()),
            title="Italic",
            input_message_content=InputTextMessageContent(
                f"<i>{escape(query)}</i>", parse_mode=ParseMode.HTML
            ),
        ),
    ]

    await update.inline_query.answer(results)


async def msnv(update: Update, context: CallbackContext) -> None:
    args = context.args
    des_msnv = "Không tìm thấy mã số nhân viên"

    if len(args) == 1:
        clone = args[0]
        name = ""
        with lite.connect(database_path) as conn:
            cursor = conn.cursor()
            str = remove_sign_for_vietnamese_string(clone)
            cursor.execute("SELECT * FROM Employees WHERE lower(Code) = ?", (str.lower(),))
            any_code = cursor.fetchone()
            if any_code is None:
                await update.message.reply_text(des_msnv)
                return
            name = remove_sign_for_vietnamese_string(any_code[1]).replace(" ", "_")

        await update.message.reply_chat_action(action=ChatAction.UPLOAD_PHOTO)
        des_msnv = f"Mã số nhân viên {clone}"
        _message = update.message

        stream_file = load_files_drive(f"{clone.upper()}.pdf", "Cham_Cong")
        file_image = convert_pdf_to_img(stream_file)

        if stream_file:
          await update.message.reply_document(document=file_image, filename=f"{name}.jpg")
        else:
           await update.message.reply_text("Gửi thất bại")
    else:
        await update.message.reply_text(des_msnv)


def excel_to_db():
    mess = "Update Dữ liệu thất bại"
    try:
        conn = lite.connect(database_path)
        cursor = conn.cursor()

        # Delete all records from the Employee table
        cursor.execute("DELETE FROM Employees")

        dt = excel_package_to_data_table()
        data = convert_data_table_to_list(dt, Employee)
        codes = [row.Code for row in data]

        # Check if any data with the same codes already exist in the database
        cursor.execute("SELECT Code FROM Employees WHERE Code IN ({})".format(','.join(['?']*len(codes))), codes)
        existing_codes = set(row[0] for row in cursor.fetchall())

        # Remove data with codes that already exist in the database
        data = [row for row in data if row.Code not in existing_codes]

        if data:
            # Insert new data into the Employee table
            cursor.executemany("INSERT INTO Employees (Code, EmployeeFullName, EmployeeOnlyName) VALUES (?, ?, ?)", [(row.Code, row.EmployeeFullName, row.EmployeeOnlyName) for row in data])
            conn.commit()
            mess = "Update Dữ liệu Thành công"

    except Exception as ex:
        mess = str(ex) + database_path
    finally:
        conn.close()

    return mess

def convert_data_table_to_list(dt, data_class):
    data = [get_item(data_class, row) for row in dt.to_dict(orient='records')]
    return data

def excel_package_to_data_table():
    # Load the Excel workbook
    xlFile = load_files_drive("MSNV.xlsx", "MSNV")
    workbook = load_workbook(filename=xlFile, read_only=True)

    # Select the first worksheet
    ws1 = workbook.active

    # Initialize a list to store rows
    rows = []

    for row in ws1.iter_rows(min_row=2, values_only=True):
        # Check if the first cell in the row is not empty
        if row[0]:
            # Process the row and remove Vietnamese signs
            processed_row = [remove_sign_for_vietnamese_string(cell) for cell in row]
            rows.append(processed_row)

    # Create a DataFrame from the processed rows
    df = pd.DataFrame(rows, columns=["Code", "EmployeeFullName", "EmployeeOnlyName"])

    return df

def get_item(data_class, dr):
    obj = data_class()

    for column_name in dr.keys():
        if hasattr(obj, column_name):
            setattr(obj, column_name, dr[column_name])

    return obj

def get_credentials():
    try:
        scopes = ["https://www.googleapis.com/auth/drive"]  # Full access

        key_file_path = os.path.dirname(__file__) + "\\key-drive.json"  # Path to your service account key file
        service_account_email = "data-test@my-project-upload-file-363503.iam.gserviceaccount.com"  # Your service account email

        # flow = InstalledAppFlow.from_client_secrets_file(key_file_path,scopes)
        # self.creds = flow.run_local_server(port=8080)
        # return self.creds
        credentials =  service_account.Credentials.from_service_account_file(filename=key_file_path)

        scoped_credentials = credentials.with_scopes(scopes)

        return scoped_credentials
    except Exception as e:
        print("GetCredentials: ", e)
        return None
    

def convert_pdf_to_img(pdf_stream):
    try:
       
      
        pdf_document = fitz.open(stream=pdf_stream, filetype="pdf")
        pdf_page = pdf_document.load_page(0)

        # image = pdf_page.get_pixmap(matrix=fitz.Matrix(300/72, 300/72))
        # image_bytes = image.get_image_data(output="jpeg")
        image = pdf_page.get_pixmap(matrix=fitz.Matrix(300/72, 300/72))

        pil_image = Image.frombytes("RGB", [image.width, image.height], image.samples)

        # Save the PIL image as a JPEG
        with io.BytesIO() as output:
            pil_image.save(output, format="JPEG")
            image_data = output.getvalue()

        return image_data    
        # images = pdf_page.extractText()  # Extract text from the PDF page

        if images:
            # Save the first image as a JPEG (you can save multiple images if needed)
            img = images[0]
            with io.BytesIO() as img_stream:
                img.save(img_stream, 'JPEG')
                return img_stream.getvalue()
        return None
    except Exception as ex:
        print("ConvertPdfToImg: ", ex)
        return None

def load_files_drive(file_name, folder_name):
    try:
        folder_id = get_folder(folder_name)
        credentials = get_credentials()

        drive_service = build("drive", "v3", credentials=credentials)

        query = f"'{folder_id}' in parents and name = '{file_name}'"
        results = drive_service.files().list(q=query).execute()

        file_id = results.get("files", [])[0]["id"] if "files" in results else None

        if file_id:
            request = drive_service.files().get_media(fileId=file_id)
            stream = io.BytesIO()
            downloader = MediaIoBaseDownload(stream, request)

            done = False
            while not done:
                status, done = downloader.next_chunk()
                if status:
                    print("Download %d%%." % int(status.progress() * 100))

            stream.seek(0)
            return stream

        return None
    except Exception as e:
        print("Google file: ", e)
        return None

def get_folder(name):
    try:
        credentials = get_credentials()

        drive_service = build("drive", "v3", credentials=credentials)

        query = f"name = '{name}' and mimeType = 'application/vnd.google-apps.folder'"
        results = drive_service.files().list(q=query).execute()

        folder_id = results.get("files", [])[0]["id"] if "files" in results else None

        return folder_id
    except Exception as e:
        print("Google folder: ", e)
        return None

def main() -> None:
    """Run the bot."""
    # Create the Application and pass it your bot's token.
    application = Application.builder().token("6134500832:AAGzjGW7R32GweF8KwTrH6wmIH4S0jL8rb0").build()

    # on different commands - answer in Telegram
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("msnv", msnv))
    application.add_handler(CommandHandler("adminCommand", admin_command))
    # on inline queries - show corresponding inline results
    application.add_handler(InlineQueryHandler(inline_query))

    # Run the bot until the user presses Ctrl-C
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()